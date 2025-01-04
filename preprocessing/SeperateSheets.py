from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import os

def split_excel_sheets_with_formatting(input_file, output_directory):
    # Load the workbook using openpyxl
    try:
        workbook = load_workbook(input_file)
    except Exception as e:
        print(f"Error loading file: {e}")
        return
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    # Iterate through each sheet and save it as a separate Excel file
    for sheet_name in workbook.sheetnames:
        try:
            # Create a new workbook for each sheet
            new_workbook = Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = sheet_name
            
            # Get the original sheet
            original_sheet = workbook[sheet_name]
            
            # Copy data and formatting
            for row in original_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    
                    # Copy cell formatting if possible
                    try:
                        if cell.has_style:
                            new_cell.font = cell.font
                            new_cell.border = cell.border
                            new_cell.fill = cell.fill
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection
                            new_cell.alignment = cell.alignment
                    except TypeError:
                        print(f"Warning: Could not copy style for cell {cell.coordinate} in sheet '{sheet_name}'.")
            
            # Save the new workbook
            output_file = os.path.join(output_directory, f"{sheet_name}.xlsx")
            new_workbook.save(output_file)
            print(f"Sheet '{sheet_name}' saved as '{output_file}'")
        except Exception as e:
            print(f"Error processing sheet '{sheet_name}': {e}")
    
    print("All sheets have been separated and saved with formatting preserved.")
# Usage
input_excel_file = 'wishHistory.xlsx'
output_folder = 'C:/Users/srija/OneDrive/Documents/Programming/projects/Genshin Gamble Assist/GamblingAssistGenshin/preprocessing'
split_excel_sheets_with_formatting(input_excel_file, output_folder)