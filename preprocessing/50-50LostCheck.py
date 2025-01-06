from openpyxl import load_workbook
import os

# Define the file paths
base_dir = os.path.dirname(os.path.abspath(__file__))
standard_pool_path = os.path.join(base_dir, 'StandardPool.xlsx')
character_event_path = os.path.join(base_dir, 'Character Event.xlsx')

# Load workbooks
standard_pool_wb = load_workbook(standard_pool_path)
character_event_wb = load_workbook(character_event_path)

# Load sheets
standard_pool_ws = standard_pool_wb.active
character_event_ws = character_event_wb.active

# Create a set of values from column B of StandardPool.xlsx
standard_pool_set = set()
for row in range(2, standard_pool_ws.max_row + 1):
    value = standard_pool_ws.cell(row=row, column=2).value
    if value:
        standard_pool_set.add(value)

# Add new column header "Lost 50-50" in column J
character_event_ws.cell(row=1, column=10, value="Lost 50-50")

# Iterate over rows in Character Event.xlsx
for row in range(2, character_event_ws.max_row + 1):
    value_in_d = character_event_ws.cell(row=row, column=4).value
    value_in_b = character_event_ws.cell(row=row, column=2).value
    
    # Debug print statements
    print(f"Row {row}: value_in_c = {value_in_d}, value_in_b = {value_in_b}")
    
    # Check conditions and set value in column J
    if value_in_d == 5 and value_in_b in standard_pool_set:
        print(f"Marking row {row} as Lost 50-50")
        character_event_ws.cell(row=row, column=10, value=1)
    else:
        character_event_ws.cell(row=row, column=10, value=0)

# Save the updated workbook
character_event_wb.save(character_event_path)